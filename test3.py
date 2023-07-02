from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
import json
import os
import openpyxl as xl
def get_no_of_rows(worksheet):
    count=0
    for k in range(1,(worksheet.max_row)+1):
        if(worksheet.cell(row=k,column=1).value == None):
            count=count+1
    no_of_rows = (worksheet.max_row) - count
    return no_of_rows
def get_no_of_columns(worksheet):
    countc=0
    for l in range(1,(worksheet.max_column)+1):
        if(worksheet.cell(row=1,column=l).value == None):
            countc=countc+1
    no_of_cols = (worksheet.max_column) - countc
    return no_of_cols
wb = xl.load_workbook("D:\\OLD DATA\\python")
ws=wb.active
rows = get_no_of_rows(ws)
def getlatlng(id):
    for i in range(1,rows+1):
        if(ws.cell(row=i,column=5).value == id):
            lat=ws.cell(row=i,column=8).value
            lng=ws.cell(row=i,column=9).value
            latlng=str(lat)+","+str(lng)
            return latlng
timeNumber = input("Enter Number - ")
dayhalf = input("Enter AM or PM - ")
Date = input("Enter Date in Format Month/DD/last 2 digits of year for Ex 7/27/22 - ")
for i in range(60):
    if(i<=9):
        ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+"0"+str(i)+" "+dayhalf
        wb.save("nestofix-attendance.xlsx")
    else:
        ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+str(i)+" "+dayhalf
        wb.save("nestofix-attendance.xlsx")
k=0
driver = webdriver.Chrome("D:\\OLD DATA\\Selenium Drivers\\chromedriver_win32\\chromedriver.exe")
driver.get("https://www.google.com/maps/@20.9880135,82.7525294,5z")
time.sleep(2)
elem7=driver.find_element(By.XPATH,'//*[@id="hArJGc"]') # to enter our location and destination
time.sleep(2)
elem7.click()
for ids in range(2,get_no_of_rows(ws)+1):
    idname = ws.cell(row=ids,column=5).value
    json_file = open("/mnt/c/Users/surya_h3yma/Desktop/Jsonfolder/"+idname+".json")
    print("json loaded")
    json_data_dict = json.load(json_file)
    list_of_keys=list(json_data_dict.keys())
    text1 = json_data_dict[list_of_keys[0]]
    print("First key of json ",text1)
    # time.sleep(2)
    if("AM" in text1 or "PM" in text1):
        for i in range(60):
            if(i<=9):
                ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+"0"+str(i)+" "+dayhalf
                wb.save("nestofix-attendance.xlsx")
            else:
                ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+str(i)+" "+dayhalf
                wb.save("nestofix-attendance.xlsx")
        for i in range(60):
            if(i<=9):
                for elem in range(len(list_of_keys)):
                    if Date+" "+timeNumber+":0"+str(i)+" "+dayhalf+" "+"[MY_SERVICE]:Time:" in json_data_dict[list_of_keys[elem]]:
                        text = json_data_dict[list_of_keys[elem]]
                        c =text.find("LatLng")
                        latlng=text[c+8:c+28]
                        time.sleep(1)
                        elem2=driver.find_element(By.XPATH,'//*[@id="sb_ifc51"]/input')  # start point
                        time.sleep(2)
                        latlng1=getlatlng(idname)
                        elem2.clear()
                        elem2.send_keys(latlng1)
                        time.sleep(2)
                        elem3=driver.find_element(By.XPATH,'//*[@id="sb_ifc52"]/input') # destination
                        time.sleep(2)
                        elem3.clear()
                        elem3.send_keys(latlng)
                        time.sleep(2)
                        elem5 = driver.find_element(By.XPATH,'//*[@id="directions-searchbox-1"]/button[1]') #  click search option
                        elem5.click()
                        time.sleep(2)
                        elems=driver.find_elements(By.CSS_SELECTOR,'div.ivN21e')
                        time.sleep(2)
                        print(len(elems))
                        elem6=elems[0]
                        innertext = elem6.get_attribute('innerText')
                        print(innertext.strip()) # distance in km
                        for j in range(16,get_no_of_columns(ws)):
                            if ws.cell(row=1,column=j).value in text:
                                ws.cell(row=2+k,column=j).value = innertext.strip()
                                wb.save("nestofix-attendance.xlsx")        
            else:
                for elem in range(len(list_of_keys)):
                    if Date+" "+timeNumber+":0"+str(i)+" "+dayhalf+" "+"[MY_SERVICE]:Time:" in json_data_dict[list_of_keys[elem]]:
                        text = json_data_dict[list_of_keys[elem]]
                        c =text.find("LatLng")
                        latlng=text[c+8:c+28]
                        time.sleep(1)
                        elem2=driver.find_element(By.XPATH,'//*[@id="sb_ifc51"]/input')  # start point
                        # time.sleep(2)
                        latlng1=getlatlng(idname)
                        elem2.clear()
                        elem2.send_keys(latlng1)
                        time.sleep(2)
                        elem3=driver.find_element(By.XPATH,'//*[@id="sb_ifc52"]/input') # destination
                        time.sleep(2)
                        elem3.clear()
                        elem3.send_keys(latlng)
                        time.sleep(1)
                        elem5 = driver.find_element(By.XPATH,'//*[@id="directions-searchbox-1"]/button[1]') #  click search option
                        elem5.click()
                        time.sleep(2)
                        elems=driver.find_elements(By.CSS_SELECTOR,'div.ivN21e')
                        time.sleep(1)
                        print(len(elems))
                        elem6=elems[0]
                        innertext = elem6.get_attribute('innerText')
                        print(innertext.strip()) # distance in km or m
                        for j in range(16,get_no_of_columns(ws)):
                            if ws.cell(row=1,column=j).value in text:
                                ws.cell(row=2+k,column=j).value = innertext.strip()# will give in distance
                                wb.save("nestofix-attendance.xlsx")
    else:
        if(dayhalf=="PM"):
            timeNumber=str(int(timeNumber)+12)
        for i in range(60):
            if(i<=9):
                ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+"0"+str(i)
                wb.save("nestofix-attendance.xlsx")
            else:
                ws.cell(row=1,column=16+i).value=str(timeNumber)+":"+str(i)
                wb.save("nestofix-attendance.xlsx")
        for i in range(60):
            if(i<=9):
                for elem in range(len(list_of_keys)):
                    if Date+" "+timeNumber+":0"+str(i)+" "+"[MY_SERVICE]:Time:" in json_data_dict[list_of_keys[elem]]:
                        text = json_data_dict[list_of_keys[elem]]
                        c =text.find("LatLng")
                        latlng=text[c+8:c+28]
                        time.sleep(1)
                        elem2=driver.find_element(By.XPATH,'//*[@id="sb_ifc51"]/input')  # start point
                        time.sleep(2)
                        latlng1=getlatlng(idname)
                        elem2.clear()
                        elem2.send_keys(latlng1)
                        time.sleep(2)
                        elem3=driver.find_element(By.XPATH,'//*[@id="sb_ifc52"]/input') # destination
                        time.sleep(2)
                        elem3.clear()
                        elem3.send_keys(latlng)
                        time.sleep(1)
                        elem5 = driver.find_element(By.XPATH,'//*[@id="directions-searchbox-1"]/button[1]') #  click search option
                        elem5.click()
                        time.sleep(2)
                        elems=driver.find_elements(By.CSS_SELECTOR,'div.ivN21e')
                        time.sleep(1)
                        print(len(elems))
                        elem6=elems[0]
                        innertext = elem6.get_attribute('innerText')
                        print(innertext.strip()) # distance in km
                        for j in range(16,get_no_of_columns(ws)):
                            if ws.cell(row=1,column=j).value in text:
                                ws.cell(row=2+k,column=j).value = innertext.strip()
                                wb.save("nestofix-attendance.xlsx")        
            else:
                for elem in range(len(list_of_keys)):
                    if Date+" "+timeNumber+":0"+str(i)+" "+"[MY_SERVICE]:Time:" in json_data_dict[list_of_keys[elem]]:
                        text = json_data_dict[list_of_keys[elem]]
                        c =text.find("LatLng")
                        latlng=text[c+8:c+28]
                        time.sleep(1)
                        elem2=driver.find_element(By.XPATH,'//*[@id="sb_ifc51"]/input')  # start point
                        time.sleep(2)
                        latlng1=getlatlng(idname)
                        elem2.clear()
                        elem2.send_keys(latlng1)
                        time.sleep(2)
                        elem3=driver.find_element(By.XPATH,'//*[@id="sb_ifc52"]/input') # destination
                        time.sleep(2)
                        elem3.clear()
                        elem3.send_keys(latlng)
                        time.sleep(1)
                        elem5 = driver.find_element(By.XPATH,'//*[@id="directions-searchbox-1"]/button[1]') #  click search option
                        elem5.click()
                        time.sleep(2)
                        elems=driver.find_elements(By.CSS_SELECTOR,'div.ivN21e')
                        time.sleep(1)
                        print(len(elems))
                        elem6=elems[0]
                        innertext = elem6.get_attribute('innerText')
                        print(innertext.strip()) # distance in km or m
                        for j in range(16,get_no_of_columns(ws)):
                            if ws.cell(row=1,column=j).value in text:
                                ws.cell(row=2+k,column=j).value = innertext.strip()# will give in distance
                                wb.save("nestofix-attendance.xlsx")


    k=k+1
#//*[@id="section-directions-trip-0"]/div[1]/div[3]/div[1]/div[2]
#//*[@id="section-directions-trip-0"]/div[1]/div[1]/div[1]/div[2]/div